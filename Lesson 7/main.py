import json, math
import openpyxl

def distance(city_1,city_2):                                # Функция с прошлого домашнего задания
        # Перевод координат в радиан
        lat_1 = float(city_1[0]) * math.pi/180              #!          
        lng_1 = float(city_1[1]) * math.pi/180              #!  Почему пришлось переводить в Float значения списка координат
        lat_2 = float(city_2[0]) * math.pi/180              #!
        lng_2 = float(city_2[1]) * math.pi/180              #!
        # по формуле Гаверсинуса
        result_dist = 2 * 6371 * math.asin(math.sqrt(math.sin((lat_2 - lat_1)/2)**2 + \
                    math.cos(lat_1) * math.cos(lat_2) * math.sin((lng_2 - lng_1)/2)**2))
        return result_dist


new_list_cities, title_tab_col = [],[]             # создаем список new_list_cities для преобразования в нужный словарь для доступа по ключам - названиям городов
                                                # создаем список из названий городов для заполнения заглавных столбцов и строк
                                                # и для использования как ключи для доступа к координатам

with open(r"C:/Users/USER\Documents/Python project/Git/My Lessons/Lesson 7/ua_cities.json", encoding = 'utf-8') as file:
    data = json.load(file)
    # print(type(data))
   
# Распаковываем и готовим данные из JSON файла к дальнейшей обработке и заполняем новые коллекции
    for region in (data[0]['regions']):            # уровень регионы украины
        # print(type(region))
        for cities in region['cities']:            # последний уровень - словари = имя города, координаты по меридиану
            new_list_cities.append([cities['name'], [cities['lat'], cities['lng']]])
            title_tab_col.append(cities['name'])
my_dict = dict(new_list_cities)                    # Преобразовываем список в словарь для доступа по ключу
print(new_list_cities)

print("Список городов для определения расстояния:  " + str(title_tab_col))

# Блок записи в файл
workbook = openpyxl.load_workbook(r'C:\Users\USER\Documents\Python project\Git\My Lessons\Lesson 7\out_file.xlsx')
sheet = workbook['data']                                      # # Выбираем лист, с которым будем работать

# Создаем столбики с названиями городов в файле
for i in range(len(title_tab_col)):
    sheet.cell(row=1, column=i+2).value = title_tab_col[i]

# Создаем строчки в файле
for i in range(len(title_tab_col)):
    sheet.cell(row=i+2, column=1).value = title_tab_col[i]
    for x in range(len(title_tab_col)):
         sheet.cell(row=i+2, column=x+2).value = distance((my_dict[title_tab_col[x]]),(my_dict[title_tab_col[i]]))

workbook.save(r'C:\Users\USER\Documents\Python project\Git\My Lessons\Lesson 7\out_file.xlsx')



# Блок рассчета расстояния между городами по функции
city_1 = input("Ввведите город отправления: ")
city_2 = input("Ввведите город Прибытия: ")

try:
    print(f"Используя собственную функцию по формуле Гаверсинуса: {distance((my_dict[city_1]),(my_dict[city_2])):.2f} километров")
except:
    print("Первый или второй город не найден в словаре")

